from flask import Flask, render_template, jsonify, request, Response, send_file, session, redirect, url_for
import csv, io, os
from datetime import date
import database as db

app = Flask(__name__)
app.secret_key = "edutrack-secret-key-2026-secure"

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "static", "photos")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db.init_db()

# ── Auth & Protection ──────────────────────────────────────────────────────────

@app.before_request
def require_login():
    public_paths = ["/login", "/static/"]
    if any(request.path.startswith(p) for p in public_paths):
        return None
    if "user_id" not in session:
        if request.path.startswith("/api/"):
            return jsonify({"error": "Unauthorized", "redirect": "/login"}), 401
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session and request.method == "GET":
        return redirect(url_for("index"))
    
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = db.verify_user(username, password)
        if user:
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["user_name"] = user["name"]
            session["user_role"] = user["role"]
            db.log_audit("LOGIN", "User", f"User logged in: {user['name']} ({user['role']})")
            return redirect(url_for("index"))
        return render_template("login.html", error="Invalid username or password.", username=username)

    return render_template("login.html")


@app.route("/logout")
def logout():
    name = session.get("user_name", "User")
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/me")
def api_me():
    if "user_id" not in session:
        return jsonify({"error": "Not logged in"}), 401
    return jsonify({
        "id": session.get("user_id"),
        "username": session.get("username"),
        "name": session.get("user_name"),
        "role": session.get("user_role")
    })


# ── Pages ──────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", user={
        "name": session.get("user_name", "User"),
        "role": session.get("user_role", "Teacher"),
        "username": session.get("username", "")
    })

# ── API: Dashboard ─────────────────────────────────────────────────────────────

@app.route("/api/dashboard")
def api_dashboard():
    return jsonify(db.get_dashboard_stats())

@app.route("/api/stats/monthly")
def api_monthly_stats():
    months = request.args.get("months", 6, type=int)
    return jsonify(db.get_monthly_stats(months))

@app.route("/api/stats/rankings")
def api_rankings():
    class_id = request.args.get("class_id", type=int)
    limit = request.args.get("limit", 10, type=int)
    return jsonify(db.get_rankings(class_id, limit))

# ── API: Classes ───────────────────────────────────────────────────────────────

@app.route("/api/classes", methods=["GET", "POST"])
def api_classes():
    if request.method == "GET":
        return jsonify(db.get_all_classes())
    data = request.json
    db.add_class(data["name"], data.get("section", ""))
    return jsonify({"ok": True})

@app.route("/api/classes/<int:class_id>", methods=["DELETE"])
def api_delete_class(class_id):
    db.delete_class(class_id)
    return jsonify({"ok": True})

# ── API: Subjects ──────────────────────────────────────────────────────────────

@app.route("/api/subjects", methods=["GET", "POST"])
def api_subjects():
    if request.method == "GET":
        class_id = request.args.get("class_id", type=int)
        return jsonify(db.get_all_subjects(class_id))
    data = request.json
    db.add_subject(data["name"], data.get("class_id"), data.get("code", ""))
    return jsonify({"ok": True})

@app.route("/api/subjects/<int:subject_id>", methods=["DELETE"])
def api_delete_subject(subject_id):
    db.delete_subject(subject_id)
    return jsonify({"ok": True})

@app.route("/api/subjects/<int:subject_id>/attendance")
def api_subject_attendance(subject_id):
    target_date = request.args.get("date", date.today().isoformat())
    class_id = request.args.get("class_id", type=int)
    return jsonify(db.get_subject_attendance_for_date(subject_id, target_date, class_id))

@app.route("/api/subjects/<int:subject_id>/attendance/mark", methods=["POST"])
def api_mark_subject_attendance(subject_id):
    data = request.json
    target_date = data.get("date", date.today().isoformat())
    records = data.get("records", [])
    db.mark_bulk_subject_attendance(records, subject_id, target_date)
    return jsonify({"ok": True, "marked": len(records)})

# ── API: Students ──────────────────────────────────────────────────────────────

@app.route("/api/students", methods=["GET", "POST"])
def api_students():
    if request.method == "GET":
        class_id = request.args.get("class_id", type=int)
        return jsonify(db.get_all_students(class_id))
    data = request.json
    try:
        db.add_student(
            data["roll_number"], data["name"],
            data.get("class_id"), data.get("email", ""), data.get("phone", "")
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/students/<int:student_id>", methods=["GET", "PUT", "DELETE"])
def api_student(student_id):
    if request.method == "GET":
        student = db.get_student_by_id(student_id)
        if not student:
            return jsonify({"error": "Not found"}), 404
        student["stats"] = db.get_student_stats(student_id)
        student["history"] = db.get_student_attendance_history(student_id, 90)
        student["subject_stats"] = db.get_student_subject_stats(student_id)
        return jsonify(student)
    if request.method == "PUT":
        data = request.json
        db.update_student(
            student_id, data["roll_number"], data["name"],
            data.get("class_id"), data.get("email", ""), data.get("phone", "")
        )
        return jsonify({"ok": True})
    db.delete_student(student_id)
    return jsonify({"ok": True})

@app.route("/api/students/search")
def api_search_students():
    q = request.args.get("q", "")
    return jsonify(db.search_students(q))

@app.route("/api/students/import", methods=["POST"])
def api_bulk_import():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    stream = io.TextIOWrapper(file.stream, encoding="utf-8-sig")
    reader = csv.DictReader(stream)
    rows = [r for r in reader]
    success, failed = db.bulk_import_students(rows)
    return jsonify({"ok": True, "success": success, "failed": failed, "total": len(rows)})

@app.route("/api/students/<int:student_id>/photo", methods=["POST"])
def api_upload_photo(student_id):
    file = request.files.get("photo")
    if not file:
        return jsonify({"error": "No file"}), 400
    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "webp"):
        return jsonify({"error": "Invalid file type"}), 400
    filename = f"student_{student_id}.{ext}"
    path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(path)
    db.update_student_photo(student_id, f"/static/photos/{filename}")
    return jsonify({"ok": True, "photo": f"/static/photos/{filename}"})

# ── API: Attendance ────────────────────────────────────────────────────────────

@app.route("/api/attendance")
def api_attendance():
    target_date = request.args.get("date", date.today().isoformat())
    class_id = request.args.get("class_id", type=int)
    return jsonify(db.get_attendance_for_date(target_date, class_id))

@app.route("/api/attendance/mark", methods=["POST"])
def api_mark_attendance():
    data = request.json
    target_date = data.get("date", date.today().isoformat())
    records = data.get("records", [])
    db.mark_bulk_attendance(records, target_date)
    return jsonify({"ok": True, "marked": len(records)})

@app.route("/api/attendance/mark-single", methods=["POST"])
def api_mark_single():
    data = request.json
    db.mark_attendance(data["student_id"], data["date"], data["status"], data.get("notes", ""))
    return jsonify({"ok": True})

# ── API: Holidays ──────────────────────────────────────────────────────────────

@app.route("/api/holidays", methods=["GET", "POST"])
def api_holidays():
    if request.method == "GET":
        return jsonify(db.get_all_holidays())
    data = request.json
    db.add_holiday(data["date"], data["name"], data.get("type", "Holiday"))
    return jsonify({"ok": True})

@app.route("/api/holidays/<int:holiday_id>", methods=["DELETE"])
def api_delete_holiday(holiday_id):
    db.delete_holiday(holiday_id)
    return jsonify({"ok": True})

@app.route("/api/holidays/upcoming")
def api_upcoming_holidays():
    return jsonify(db.get_upcoming_holidays(5))

# ── API: Leaves ────────────────────────────────────────────────────────────────

@app.route("/api/leaves", methods=["GET", "POST"])
def api_leaves():
    if request.method == "GET":
        return jsonify(db.get_all_leaves())
    data = request.json
    db.add_leave(
        data["student_id"], data["from_date"], data["to_date"],
        data.get("reason", ""), data.get("status", "Approved")
    )
    return jsonify({"ok": True})

@app.route("/api/leaves/<int:leave_id>", methods=["PUT", "DELETE"])
def api_leave(leave_id):
    if request.method == "PUT":
        data = request.json
        db.update_leave_status(leave_id, data["status"])
        return jsonify({"ok": True})
    db.delete_leave(leave_id)
    return jsonify({"ok": True})

@app.route("/api/leaves/apply", methods=["POST"])
def api_apply_leaves():
    count = db.apply_approved_leaves()
    return jsonify({"ok": True, "updated": count})

# ── API: Audit Log ─────────────────────────────────────────────────────────────

@app.route("/api/audit")
def api_audit():
    limit = request.args.get("limit", 100, type=int)
    return jsonify(db.get_audit_log(limit))

@app.route("/api/audit/clear", methods=["POST"])
def api_clear_audit():
    db.clear_audit_log()
    return jsonify({"ok": True})

# ── API: Reports ───────────────────────────────────────────────────────────────

@app.route("/api/reports")
def api_reports():
    start = request.args.get("start", date.today().replace(day=1).isoformat())
    end   = request.args.get("end", date.today().isoformat())
    class_id = request.args.get("class_id", type=int)
    return jsonify(db.get_attendance_report(start, end, class_id))

@app.route("/api/reports/export")
def api_export_csv():
    start = request.args.get("start", date.today().replace(day=1).isoformat())
    end   = request.args.get("end", date.today().isoformat())
    class_id = request.args.get("class_id", type=int)
    rows = db.get_attendance_report(start, end, class_id)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "roll_number", "name", "class_name",
        "total_days", "present", "absent", "late", "excused", "percentage"
    ])
    writer.writeheader()
    writer.writerows(rows)
    return Response(
        output.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=attendance_{start}_to_{end}.csv"}
    )

@app.route("/api/reports/export-excel")
def api_export_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    start = request.args.get("start", date.today().replace(day=1).isoformat())
    end   = request.args.get("end", date.today().isoformat())
    class_id = request.args.get("class_id", type=int)
    rows = db.get_attendance_report(start, end, class_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_fill = PatternFill("solid", fgColor="1a2332")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    good_fill   = PatternFill("solid", fgColor="c6efce")
    warn_fill   = PatternFill("solid", fgColor="ffeb9c")
    danger_fill = PatternFill("solid", fgColor="ffc7ce")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="d0d7de")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Roll No.", "Student Name", "Class",
               "Total Days", "Present", "Absent", "Late", "Excused", "Attendance %"]
    col_widths = [5, 12, 28, 24, 11, 10, 10, 10, 10, 14]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(rows, 2):
        pct = row["percentage"]
        fill = good_fill if pct >= 75 else (warn_fill if pct >= 50 else danger_fill)
        vals = [ri - 1, row["roll_number"], row["name"], row.get("class_name",""),
                row["total_days"], row["present"], row["absent"],
                row["late"], row["excused"], f"{pct}%"]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center
            cell.border = border
            if ci == 10:
                cell.fill = fill
                cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{start}_to_{end}.xlsx"}
    )

# ── API: QR Code ───────────────────────────────────────────────────────────────

@app.route("/api/students/<int:student_id>/qr")
def api_qr_code(student_id):
    import qrcode
    student = db.get_student_by_id(student_id)
    if not student:
        return jsonify({"error": "Not found"}), 404
    data = f"ID:{student_id}|Roll:{student['roll_number']}|Name:{student['name']}|Class:{student.get('class_name','')}"
    qr = qrcode.QRCode(version=1, box_size=8, border=3)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#0d1117", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")

# ── API: Backup ────────────────────────────────────────────────────────────────

@app.route("/api/backup")
def api_backup():
    import database as db2
    backup_path = db2.DB_PATH
    if not os.path.exists(backup_path):
        return jsonify({"error": "Database not found"}), 404
    today = date.today().isoformat()
    return send_file(
        backup_path,
        as_attachment=True,
        download_name=f"edutrack_backup_{today}.db",
        mimetype="application/octet-stream"
    )

# ── API: CSV Template ──────────────────────────────────────────────────────────

@app.route("/api/students/import/template")
def api_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["roll_number", "name", "class_name", "email", "phone"])
    writer.writerow(["1001", "Sample Student", "B.Tech CSE · Sem 1 - Section A", "student@school.edu", "9876543210"])
    return Response(
        output.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=students_import_template.csv"}
    )


if __name__ == "__main__":
    print("\n  Student Attendance System is running!")
    print("   Open your browser at -> http://127.0.0.1:5000\n")
    app.run(host="127.0.0.1", port=5000, debug=True, use_reloader=False)
